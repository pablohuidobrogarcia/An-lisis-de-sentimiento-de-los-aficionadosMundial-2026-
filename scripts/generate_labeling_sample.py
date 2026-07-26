from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.protection import Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

RANDOM_STATE_SAMPLE = 20260718
RANDOM_STATE_SPLIT = 42
BASE = Path(
    r"C:\Users\Pablo\Desktop\Proyecto Mundial\An-lisis-de-sentimiento-de-los-aficionadosMundial-2026-"
)

DATA = (
    BASE
    / "data"
    / "processed"
    / "comentarios_sentimiento"
    / "comentarios_sentimiento.parquet"
)
EXISTING_LABELS = BASE / "evaluation" / "manual_labels_random_sample.csv"
OUT_DIR = BASE / "data" / "eval"
OUT_DIR.mkdir(parents=True, exist_ok=True)

df = pd.read_parquet(DATA)
print(f"Cargadas {len(df)} filas, {len(df.columns)} columnas")

dup_ids = [
    "UgzErRUyWp1Ui89t7J54AaABAg",
    "Ugx7OivWxhQR1sLilCV4AaABAg.AYGfoePxbDPAYHViX24gas",
]
before = len(df)
df = df[~df["comment_id"].isin(dup_ids)].reset_index(drop=True)
print(f"Excluidos {before - len(df)} rows de comment_id duplicados")

existing = pd.read_csv(EXISTING_LABELS, encoding="utf-8")
existing_ids = set(existing["comment_id"].tolist())
before = len(df)
df = df[~df["comment_id"].isin(existing_ids)].reset_index(drop=True)
print(f"Excluidos {before - len(df)} rows ya etiquetados (muestra de 200)")
print(f"Pool disponible para muestreo: {len(df)} filas")

lang_dist = df["language"].value_counts()
total_pool = len(df)
print("\n=== Distribucion de idiomas (pool) ===")
for lang, count in lang_dist.items():
    print(f"  {lang}: {count:>6} ({count/total_pool*100:.1f}%)")

sample = df.sample(n=3300, random_state=RANDOM_STATE_SAMPLE).reset_index(drop=True)
print(f"\nMuestra aleatoria: {len(sample)} filas (seed={RANDOM_STATE_SAMPLE})")

lang_sample = sample["language"].value_counts()
print("=== Distribucion de idiomas (muestra) ===")
for lang, count in lang_sample.items():
    print(f"  {lang}: {count:>6} ({count/3300*100:.1f}%)")

train = sample.sample(n=2800, random_state=RANDOM_STATE_SPLIT)
test = sample.drop(train.index).reset_index(drop=True)
train = train.reset_index(drop=True)
print(f"\nTrain: {len(train)} filas  |  Test: {len(test)} filas")
assert len(train) + len(test) == 3300
assert set(train["comment_id"]).isdisjoint(set(test["comment_id"]))
assert set(train["comment_id"]).isdisjoint(existing_ids)
assert set(test["comment_id"]).isdisjoint(existing_ids)

mapping_cols = ["comment_id", "sentiment_bert", "sentiment_model_version"]
for label, split_df in [("train", train), ("test", test)]:
    mapping_path = OUT_DIR / f"model_predictions_for_finetuning_{label}.csv"
    split_df[mapping_cols].to_csv(mapping_path, index=False, encoding="utf-8")
    print(f"Mapping: {mapping_path.name} ({len(split_df)} filas)")


def build_xlsx(df_split, path, split_name):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Labeling {split_name}"

    visible_cols = ["comment_id", "text_clean", "language"]
    out = df_split[visible_cols].copy()
    out["manual_label"] = ""

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), 1):
        ws.append(list(row))
        if r_idx == 1:
            for c in range(1, len(row) + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 100
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 15

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.protection = Protection(locked=True)

    dv = DataValidation(type="list", formula1='"POS,NEU,NEG"', allow_blank=True)
    dv.error = "Solo se permiten los valores: POS, NEU, NEG"
    dv.errorTitle = "Valor invalido"
    dv.prompt = "Selecciona POS, NEU o NEG"
    dv.promptTitle = "Etiqueta de sentimiento"
    ws.add_data_validation(dv)
    dv.add(f"D2:D{ws.max_row}")

    ws.freeze_panes = "A2"

    wb.save(path)
    mb = path.stat().st_size / 1_000_000
    print(f"XLSX: {path.name} ({ws.max_row - 1} filas, {mb:.2f} MB)")


build_xlsx(train, OUT_DIR / "sentiment_labeling_train.xlsx", "Train")
build_xlsx(test, OUT_DIR / "sentiment_labeling_test.xlsx", "Test")

print("\n" + "=" * 70)
print("  RESUMEN FINAL")
print("=" * 70)
print(f"  Pool (excl. dups + 200): {total_pool} filas")
print(f"  Muestra: 3,300 (seed={RANDOM_STATE_SAMPLE})")
print(f"  Train:   2,800 / Test: 500 (seed split={RANDOM_STATE_SPLIT})")
print("  No overlap existing 200: ok")
print("  No overlap train vs test: ok")
print("  Duplicados excluidos: ok")
print()
for f in sorted(OUT_DIR.glob("sentiment_labeling_*")):
    print(f"  {f.name}  ({f.stat().st_size / 1_000_000:.2f} MB)")
for f in sorted(OUT_DIR.glob("model_predictions_for_finetuning_*")):
    print(f"  {f.name}  ({f.stat().st_size / 1_000_000:.2f} MB)")
